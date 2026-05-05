#!/usr/bin/env python3
"""
XLSX Generator for eWards Lead Generation
Takes merged lead data and produces a single-sheet XLSX with combined company + people data.

Columns: Company Name | Vertical / Nature | Website URL | Company LinkedIn | Person Name | Designation | Person LinkedIn

Usage:
    python generate_xlsx.py --input merged_leads.json --output leads.xlsx
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def create_xlsx(data, output_path):
    """Generate a single-sheet XLSX with combined company + people data."""

    companies = data.get("companies", [])
    people = data.get("people", [])

    # Build company lookup
    company_lookup = {}
    for c in companies:
        company_lookup[c.get("name", "")] = c

    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    link_font = Font(name="Calibri", size=11, color="3B82F6", underline="single")
    body_font = Font(name="Calibri", size=11)
    body_alignment = Alignment(vertical="center", wrap_text=False)
    header_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB')
    )

    ws = wb.active
    ws.title = "Leads"

    headers = [
        "Company Name", "Vertical / Nature", "Website URL",
        "Person Name", "Designation", "Person LinkedIn",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, person in enumerate(people, 2):
        company_name = person.get("company", "")
        c = company_lookup.get(company_name, {})

        # Company Name
        cell = ws.cell(row=row_idx, column=1, value=company_name)
        cell.font = body_font
        cell.alignment = body_alignment
        cell.border = thin_border

        # Vertical / Nature
        cell = ws.cell(row=row_idx, column=2, value=c.get("vertical", ""))
        cell.font = body_font
        cell.alignment = body_alignment
        cell.border = thin_border

        # Website URL
        website = c.get("website", "")
        cell = ws.cell(row=row_idx, column=3, value=website)
        if website:
            cell.font = link_font
            try:
                cell.hyperlink = website if website.startswith("http") else f"https://{website}"
            except Exception:
                pass
        cell.alignment = body_alignment
        cell.border = thin_border

        # Person Name
        cell = ws.cell(row=row_idx, column=4, value=person.get("name", ""))
        cell.font = body_font
        cell.alignment = body_alignment
        cell.border = thin_border

        # Designation
        cell = ws.cell(row=row_idx, column=5, value=person.get("designation", ""))
        cell.font = body_font
        cell.alignment = body_alignment
        cell.border = thin_border

        # Person LinkedIn
        p_li = person.get("linkedin", person.get("linkedin_url", ""))
        cell = ws.cell(row=row_idx, column=6, value=p_li)
        if p_li:
            cell.font = link_font
            try:
                cell.hyperlink = p_li
            except Exception:
                pass
        cell.alignment = body_alignment
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 45
    ws.freeze_panes = "A2"

    wb.save(output_path)

    print(f"XLSX saved: {output_path}")
    print(f"  Sheet 'Leads': {len(people)} rows")


def main():
    parser = argparse.ArgumentParser(description="Generate leads XLSX")
    parser.add_argument("--input", required=True, help="Merged JSON file")
    parser.add_argument("--output", default=None, help="Output XLSX path")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: {args.input} not found")
        sys.exit(1)

    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if not args.output:
        meta = data.get("metadata", {})
        safe_type = meta.get("type", "leads").replace(" ", "_").lower()
        safe_city = meta.get("city", "unknown").replace(" ", "_").lower()
        date_str = datetime.now().strftime("%Y%m%d")
        args.output = f"leads_{safe_type}_{safe_city}_{date_str}.xlsx"

    create_xlsx(data, args.output)


if __name__ == "__main__":
    main()
